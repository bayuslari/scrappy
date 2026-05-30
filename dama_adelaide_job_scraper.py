#!/usr/bin/env python3
"""
DAMA Adelaide Job Scraper
=========================
Scrapes frontend/fullstack developer jobs in Adelaide from Indeed, LinkedIn,
and SEEK (custom scraper — jobspy doesn't support SEEK natively).

Note: Jobstreet Australia is owned by SEEK and redirects to the same listings,
so scraping SEEK covers both.

Requirements:
    pip install python-jobspy pandas openpyxl requests

Usage:
    python dama_adelaide_job_scraper.py

Output:
    - dama_jobs_YYYYMMDD.csv   (all results)
    - dama_jobs_YYYYMMDD.xlsx  (formatted, ready to review)
"""

import re
import sys
import json
import time
import logging
from datetime import datetime
from pathlib import Path

# ── dependencies check ────────────────────────────────────────────────────────
try:
    import pandas as pd
    import requests
    from jobspy import scrape_jobs
except ImportError:
    print("❌  Missing dependencies. Run:\n    pip install python-jobspy pandas openpyxl requests")
    sys.exit(1)

# ── config ────────────────────────────────────────────────────────────────────
TODAY       = datetime.now().strftime("%Y%m%d")
OUTPUT_DIR  = Path(".")          # change to e.g. Path("~/jobs").expanduser()
LOG_LEVEL   = logging.INFO

logging.basicConfig(
    level=LOG_LEVEL,
    format="%(asctime)s  %(levelname)-8s %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("dama-scraper")

# ── search configuration ──────────────────────────────────────────────────────
# ── Customised for Bayu Riyadi — Senior Frontend Developer ───────────────────
# CV skills: React, Next.js, TypeScript, Tailwind CSS, Supabase, GraphQL,
#            Shopify (Headless), Prismic CMS, Vercel, AWS, Cypress, Jest,
#            Zustand, Framer, Shadcn/ui, React Native
# Experience: 10+ years | Fintech, SaaS, E-commerce | Remote/Relocation open
# ─────────────────────────────────────────────────────────────────────────────

# Search queries — broad enough to catch mid/senior roles alike
# (seniority is shown in the listing itself; no need to filter by title)
SEARCH_QUERIES = [
    "Frontend Developer React",
    "React Developer",
    "Frontend Engineer Next.js",
    "Frontend Engineer TypeScript",
    "Next.js developer",
    "JavaScript developer",
    "React TypeScript developer",
    "Headless CMS developer React",
    "Shopify developer React",
    "Full stack developer Next.js",
]

# Glassdoor consistently returns 400 for "Adelaide, SA" (location not parsed).
# SEEK is handled separately via custom scraper below — not supported by jobspy.
SITES = ["indeed", "linkedin"]

LOCATION      = "Adelaide, SA"
RESULTS_EACH  = 30       # per query per site (keep reasonable to avoid blocks)
MAX_AGE_DAYS  = 30       # only jobs posted within this many days
DELAY_BETWEEN = 3        # seconds between requests (be polite)

# ── sponsorship signal keywords ───────────────────────────────────────────────
# Positive signals — employer explicitly open to sponsorship / migration
# NOTE: "work rights" / "right to work" means employer is *asking* you to have
# them — NOT offering sponsorship. Strong signals are: 482, TSS, sponsorship,
# DAMA, employer sponsored, skilled migration.
SPONSOR_POS = [
    r"visa sponsor",
    r"sponsorship available",
    r"sponsorship provided",
    r"\bsponsorship\b",
    r"DAMA",
    r"skilled migration",
    r"\b482\b",
    r"TSS visa",
    r"employer.?sponsored",
    r"relocation (assistance|support|package)",
    r"open to relocation",
    r"will sponsor",
    r"willing to sponsor",
]

# Weak positive — employer mentions work rights (could go either way)
SPONSOR_WEAK = [
    r"work rights",
    r"right to work",
    r"relocation",
]

# Negative signals — employer explicitly excludes overseas applicants
SPONSOR_NEG = [
    r"must be (an )?australian citizen",
    r"australian (citizens?|residents?) only",
    r"permanent residents? only",
    r"no sponsorship",
    r"must have full.?working rights",
    r"must hold (a )?permanent",
    r"clearance required",
    r"security clearance",
    r"nv1|nv2|baseline clearance",
    r"australian citizenship (is )?required",
]

# Tech stack — matched to Bayu's CV (higher score = better fit)
TECH_KEYWORDS = [
    # Core (Bayu's primary stack — high weight)
    "react", "next.js", "nextjs", "typescript", "tailwind", "tailwindcss",
    # Strong secondary
    "javascript", "supabase", "graphql", "rest api",
    # Tooling / platforms Bayu uses
    "vercel", "aws", "shadcn", "framer", "zustand", "prismic",
    "shopify", "headless", "cms",
    # Testing / quality
    "cypress", "jest",
    # Other frameworks (present in CV)
    "react native", "postgresql", "node.js", "nodejs",
]

# ── helpers ───────────────────────────────────────────────────────────────────

def _score_text(text: str) -> dict:
    """Score a job description for sponsorship likelihood & tech stack match.

    Sponsor labels:
      ✅ Likely   — explicit strong signals (482, TSS, sponsorship, DAMA, etc.)
      ⚠️ Weak     — only soft signals (work rights / relocation mentioned)
      🔶 Unknown  — no mention either way — approach employer directly
      ❌ No       — explicit exclusion (citizen only, security clearance, etc.)
    """
    t = (text or "").lower()

    pos_hits  = [kw for kw in SPONSOR_POS  if re.search(kw, t, re.I)]
    weak_hits = [kw for kw in SPONSOR_WEAK if re.search(kw, t, re.I)]
    neg_hits  = [kw for kw in SPONSOR_NEG  if re.search(kw, t, re.I)]
    tech_hits = [kw for kw in TECH_KEYWORDS if kw in t]

    if neg_hits:
        sponsor_label = "❌ No"
    elif pos_hits:
        sponsor_label = "✅ Likely"
    elif weak_hits:
        sponsor_label = "⚠️ Weak"
    else:
        sponsor_label = "🔶 Unknown"

    return {
        "sponsor_signal":    sponsor_label,
        "sponsor_pos_hits":  ", ".join(pos_hits + weak_hits) or "",
        "sponsor_neg_hits":  ", ".join(neg_hits)  or "",
        "tech_hits":         ", ".join(tech_hits) or "",
        "tech_score":        len(tech_hits),
    }


def scrape_seek(query: str, location: str = "Adelaide SA", max_pages: int = 3) -> pd.DataFrame:
    """Custom SEEK scraper — jobspy doesn't support SEEK natively.

    SEEK owns Jobstreet Australia, so this covers both platforms.
    Scrapes SEEK's HTML search pages and parses the __NEXT_DATA__ JSON
    that Next.js embeds in every page response.
    """
    try:
        from bs4 import BeautifulSoup
    except ImportError:
        log.warning("   ↳ SEEK skipped — install beautifulsoup4: pip install beautifulsoup4")
        return pd.DataFrame()

    SEEK_SEARCH = "https://www.seek.com.au/jobs"
    HEADERS = {
        "User-Agent": (
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/124.0.0.0 Safari/537.36"
        ),
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "en-AU,en;q=0.9",
    }

    rows = []
    seen_ids = set()

    for page in range(1, max_pages + 1):
        params = {
            "keywords": query,
            "where":    location,
            "dateRange": f"{MAX_AGE_DAYS}d",
            "page":     page,
        }
        try:
            resp = requests.get(
                SEEK_SEARCH, headers=HEADERS, params=params,
                timeout=20, allow_redirects=True,
            )
            resp.raise_for_status()
        except Exception as e:
            log.warning(f"   ↳ SEEK page {page} fetch error: {e}")
            break

        # Parse __NEXT_DATA__ embedded JSON
        soup = BeautifulSoup(resp.text, "html.parser")
        nd_tag = soup.find("script", id="__NEXT_DATA__")
        if not nd_tag:
            log.warning(f"   ↳ SEEK page {page}: __NEXT_DATA__ not found (SEEK may have changed structure)")
            break

        try:
            nd = json.loads(nd_tag.string or "{}")
        except json.JSONDecodeError:
            log.warning(f"   ↳ SEEK page {page}: could not parse __NEXT_DATA__ JSON")
            break

        # Navigate to job listings — path may vary; try known locations
        page_props = nd.get("props", {}).get("pageProps", {})
        jobs = (
            page_props.get("jobsResult", {}).get("data")
            or page_props.get("results", {}).get("data")
            or page_props.get("jobs", {}).get("data")
            or page_props.get("jobListings")
            or []
        )

        if not jobs:
            log.debug(f"   ↳ SEEK page {page}: no jobs in __NEXT_DATA__ (end of results or structure changed)")
            break

        for job in jobs:
            jid = str(job.get("id", "") or job.get("jobId", ""))
            if not jid or jid in seen_ids:
                continue
            seen_ids.add(jid)

            advertiser = job.get("advertiser") or job.get("companyMeta") or {}
            work_arrangements = job.get("workArrangements") or {}
            arrangement_labels = [
                d.get("label", "")
                for d in (work_arrangements.get("data") or [])
            ]
            is_remote = any("remote" in lbl.lower() for lbl in arrangement_labels)

            rows.append({
                "site":           "seek",
                "id":             jid,
                "job_url":        f"https://www.seek.com.au/job/{jid}",
                "job_url_direct": f"https://www.seek.com.au/job/{jid}",
                "title":          job.get("title", ""),
                "company":        advertiser.get("description", "") or advertiser.get("name", ""),
                "location":       job.get("suburb", "") or job.get("area", "") or location,
                "date_posted":    (job.get("listingDate", "") or "")[:10],
                "job_type":       (job.get("workTypes") or {}).get("label", ""),
                "description":    job.get("teaser", "") or job.get("abstract", ""),
                "is_remote":      is_remote,
                "min_amount":     None,
                "max_amount":     None,
                "currency":       None,
                "job_level":      "",
                "search_query":   query,
            })

        time.sleep(DELAY_BETWEEN)

    if not rows:
        return pd.DataFrame()

    df = pd.DataFrame(rows)
    log.info(f"   ↳ SEEK: {len(df)} jobs found")
    return df


def scrape_all() -> pd.DataFrame:
    """Run all search queries across Indeed + LinkedIn + SEEK, then deduplicate."""
    frames = []

    for query in SEARCH_QUERIES:
        log.info(f"🔍  Searching: '{query}' — {LOCATION}")

        # ── Indeed + LinkedIn (via jobspy) ────────────────────────────────────
        try:
            df = scrape_jobs(
                site_name=SITES,
                search_term=query,
                location=LOCATION,
                results_wanted=RESULTS_EACH,
                hours_old=MAX_AGE_DAYS * 24,
                country_indeed="Australia",
                linkedin_fetch_description=True,
                delay=DELAY_BETWEEN,
            )
            if not df.empty:
                df["search_query"] = query
                frames.append(df)
                log.info(f"   ↳ Indeed/LinkedIn: {len(df)} jobs found")
            else:
                log.info("   ↳ Indeed/LinkedIn: 0 jobs found")
        except Exception as e:
            log.warning(f"   ↳ Indeed/LinkedIn error for '{query}': {e}")

        # ── SEEK (custom scraper) ─────────────────────────────────────────────
        seek_df = scrape_seek(query, location="Adelaide SA")
        if not seek_df.empty:
            frames.append(seek_df)

    if not frames:
        log.error("No jobs collected from any query.")
        return pd.DataFrame()

    combined = pd.concat(frames, ignore_index=True)
    log.info(f"Total before dedup: {len(combined)}")

    # Deduplicate by job_url (most reliable unique key); fall back to title+company
    if "job_url" in combined.columns:
        combined = combined.drop_duplicates(subset=["job_url"])
    elif {"title", "company"}.issubset(combined.columns):
        combined = combined.drop_duplicates(subset=["title", "company"])

    log.info(f"Total after dedup:  {len(combined)}")
    return combined


def enrich(df: pd.DataFrame) -> pd.DataFrame:
    """Add sponsorship scoring and convenience columns."""
    desc_col = next(
        (c for c in ["description", "job_description", "body"] if c in df.columns),
        None,
    )

    scores = []
    for _, row in df.iterrows():
        desc = str(row[desc_col]) if desc_col else ""
        title = str(row.get("title", ""))
        scores.append(_score_text(desc + " " + title))

    score_df = pd.DataFrame(scores)
    df = pd.concat([df.reset_index(drop=True), score_df], axis=1)

    # Clean up date column if present
    for date_col in ["date_posted", "posted_at"]:
        if date_col in df.columns:
            df[date_col] = pd.to_datetime(df[date_col], errors="coerce").dt.strftime("%Y-%m-%d")

    return df


def save(df: pd.DataFrame):
    """Save to CSV and XLSX."""
    if df.empty:
        log.warning("Nothing to save.")
        return

    # ── column order ──────────────────────────────────────────────────────────
    priority_cols = [
        "sponsor_signal", "tech_score", "tech_hits",
        "title", "company", "location",
        "date_posted", "posted_at",
        "min_amount", "max_amount", "currency",
        "job_type", "site",
        "sponsor_pos_hits", "sponsor_neg_hits",
        "job_url", "search_query",
        "description",
    ]
    existing = [c for c in priority_cols if c in df.columns]
    rest     = [c for c in df.columns if c not in existing]
    df = df[existing + rest]

    # ── sort: likely-sponsor first, then tech score ────────────────────────────
    sponsor_order = {"✅ Likely": 0, "⚠️ Weak": 1, "🔶 Unknown": 2, "❌ No": 3}
    df = df.copy()
    df["_sort"] = df["sponsor_signal"].map(sponsor_order).fillna(3)
    df = df.sort_values(["_sort", "tech_score"], ascending=[True, False]).drop(columns=["_sort"])

    csv_path  = OUTPUT_DIR / f"dama_jobs_{TODAY}.csv"
    xlsx_path = OUTPUT_DIR / f"dama_jobs_{TODAY}.xlsx"

    # CSV
    df.to_csv(csv_path, index=False)
    log.info(f"💾  CSV  → {csv_path}")

    # XLSX with basic formatting
    try:
        with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Jobs")

            ws = writer.sheets["Jobs"]
            from openpyxl.styles import PatternFill, Font, Alignment
            from openpyxl.utils import get_column_letter

            # Header row style
            header_fill = PatternFill("solid", fgColor="1E293B")
            for cell in ws[1]:
                cell.font      = Font(color="FFFFFF", bold=True)
                cell.fill      = header_fill
                cell.alignment = Alignment(wrap_text=False)

            # Row colour-coding by sponsor signal
            green  = PatternFill("solid", fgColor="D1FAE5")  # ✅ Likely
            orange = PatternFill("solid", fgColor="FED7AA")  # ⚠️ Weak
            yellow = PatternFill("solid", fgColor="FEF9C3")  # 🔶 Unknown
            red    = PatternFill("solid", fgColor="FEE2E2")  # ❌ No

            signal_col_idx = (
                df.columns.get_loc("sponsor_signal") + 1
                if "sponsor_signal" in df.columns else None
            )

            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                if signal_col_idx:
                    val = row[signal_col_idx - 1].value or ""
                    if "Likely" in val:    fill = green
                    elif "Weak" in val:    fill = orange
                    elif "No" in val:      fill = red
                    else:                  fill = yellow
                    for cell in row:
                        cell.fill = fill

            # Auto column widths (capped)
            for i, col in enumerate(df.columns, 1):
                max_len = max(
                    len(str(col)),
                    df[col].astype(str).str.len().max() if not df.empty else 0
                )
                ws.column_dimensions[get_column_letter(i)].width = min(max_len + 2, 60)

        log.info(f"📊  XLSX → {xlsx_path}")
    except Exception as e:
        log.warning(f"XLSX save failed ({e}); CSV still saved.")

    # ── summary ───────────────────────────────────────────────────────────────
    print("\n" + "═" * 60)
    print(f"  DAMA Adelaide Job Scraper — {TODAY}")
    print("═" * 60)
    print(f"  Total unique jobs found : {len(df)}")
    if "sponsor_signal" in df.columns:
        counts = df["sponsor_signal"].value_counts()
        for label, n in counts.items():
            print(f"  {label:<20} : {n}")
    print(f"\n  Files saved to: {OUTPUT_DIR.resolve()}")
    print("═" * 60 + "\n")

    # Top picks
    top = df[~df.get("sponsor_signal", pd.Series()).isin(["❌ No"])].head(10)
    if not top.empty:
        print("🏆  Top picks (open to sponsorship, best tech match):\n")
        for _, r in top.iterrows():
            title   = r.get("title", "N/A")
            company = r.get("company", "N/A")
            signal  = r.get("sponsor_signal", "")
            tech    = r.get("tech_hits", "")
            url     = r.get("job_url", "")
            print(f"  {signal}  {title} @ {company}")
            if tech:
                print(f"       Tech: {tech}")
            if url:
                print(f"       URL : {url}")
            print()


# ── main ──────────────────────────────────────────────────────────────────────

def main():
    print("\n🚀  DAMA Adelaide Job Scraper starting...\n")
    log.info(f"Sites: {SITES}")
    log.info(f"Queries: {len(SEARCH_QUERIES)} queries × {RESULTS_EACH} results each")

    df = scrape_all()

    if df.empty:
        log.error("No jobs found. Check your internet connection or try again later.")
        sys.exit(1)

    df = enrich(df)
    save(df)


if __name__ == "__main__":
    main()
